from openpyxl import *
from openpyxl.styles import PatternFill
from datetime import datetime

import string
import random


def gentime(l,start_hour,stop_hour,step_hour,start_minute,stop_minute,step_minute):
    for hour in range(start_hour,stop_hour,step_hour):
        for minute in range(start_minute,stop_minute,step_minute):
            # Format the hour and minute with leading zeros
            formatted_hour = f"{hour:02d}"
            formatted_minute = f"{minute:02d}"
            
            # Print or use the formatted hour and minute
            l.append(f"{formatted_hour}:{formatted_minute}")
            #print(f"{formatted_hour}:{formatted_minute}")
    return l

# start_hour 0,stop_hour 23(+1),step_hour 1,start_minute 0,stop_minute 30(+1),step_minute 30

def time_cal(start_hour,end_hour,start_min,end_min):
    l = []
    if end_hour == 0:
        if start_hour == 0 or start_hour > 0:
            l = gentime(l,start_hour,24,1,start_min,end_min+1,30)
            l = gentime(l,0,1,1,start_min,end_min,30)
    elif end_hour - start_hour > 0:
        l = gentime(l,start_hour,end_hour+1,1,start_min,end_min+1,30)
    else:
        l = gentime(l,start_hour,24,1,start_min,end_min+1,30)
        l = gentime(l,0,end_hour+1,1,start_min,end_min+1,30)
    return l

def gen_letter(max_letter = "AY"):
    l = []
    ascii_value = 65 
    to_Z = False
    count = 0 #max = 51 #mid = 26
    letter = ""
    while True:
        if letter == max_letter:
            to_Z = False
            break
        if ascii_value == 91:
            ascii_value = 65
            to_Z = True
        else:
            if to_Z == True:
                letter = "A"+chr(ascii_value)
                l.append(letter)
                ascii_value+=1
                count+=1
            else:
                letter = chr(ascii_value)
                l.append(letter)
                ascii_value+=1
                count+=1
    return l

def main(SheetName):
#------------------- open excel file --------------------------------
    try:
        wb = load_workbook(SheetName)
    except FileNotFoundError:
        print("Sheet not found")
#----------------------------------------------------------------

#------------------- access sheet or create sheet --------------------------------    
    if "graph" in wb.sheetnames:
        ws = wb["graph"]
    else:
        ws = wb.create_sheet("graph")
#----------------------------------------------------------------

#------------------- generate time --------------------------------
    start_hour = 0
    end_hour = 0
    #step_hour = 1
    start_min = 0
    end_min = 31
    #step_minute = 30

    timer = time_cal(start_hour,end_hour,start_min,end_min)
    insert_pos = 1
    round = len(timer)
    while True:
        if round == 0:
            break
        else:
            ws.cell(row=1, column=insert_pos+1, value=timer[insert_pos-1])
            insert_pos += 1
            round-=1
#----------------------------------------------------------------   

##------------------- generate Tools --------------------------------
    ws.cell(row=3,column=1,value="Blending")
    ws.cell(row=5,column=1,value="Circulate")
#----------------------------------------------------------------

#------------------- color matrix --------------------------------
    color = ["00FF0000","0000FF00","000000FF","00FFFF00","00FF00FF","0000FFFF",
              "00FF0000","0000FF00","000000FF","00FFFF00","00FF00FF","0000FFFF",
              "00008000","00808000","00800080","00008080","00C0C0C0","00808080",
              "009999FF","00993366","00FFFFCC","00CCFFFF","00FF8080","000066CC",
              "00CCCCFF","00FF00FF","00FFFF00","0000FFFF","00800080","00008080",
              "000000FF","0000CCFF","00CCFFFF","00CCFFCC","00FFFF99","0099CCFF",
              "00FF99CC","00CC99FF","00FFCC99","003366FF","0033CCCC","0099CC00",
              "00FFCC00","00FF9900","00FF6600","00666699","00969696","00339966",
              "00993300","00993366","00333399"]
#----------------------------------------------------------------

#------------------- Ploting in sheet --------------------------------
    wsr = wb["ABB1"]
    #letter = gen_letter()
    for r in wsr.iter_rows(min_row=2):  # Start from the second row
        use_color = random.randint(0,len(color)-1)
        round = 1
        dict ={} #Tool(Blending or Circulate), P_Name, Start, End
        for cell in r:
            if round == 1:
                dict["Tool"] = cell.value
            elif round == 2:
                dict["P_Name"] = cell.value
            elif round == 3:
                dict["Start"] = cell.value.strftime("%H:%M")
            elif round == 4:
                dict["End"] = cell.value.strftime("%H:%M")
            else:
                print("??")
            round+=1
        start_hour = timer.index(dict["Start"])+2
        end_hour = timer.index(dict["End"])+3
        Fill = PatternFill(start_color=color[use_color],end_color=color[use_color],fill_type="solid")
        for i in range(start_hour, end_hour):
            if dict["Tool"] == "Blending":
                if i == start_hour:
                    ws.cell(row=3,column=i,value=dict["P_Name"]).fill=Fill
                else:
                    ws.cell(row=3,column=i).fill=Fill
            else:
                if i == start_hour:
                    ws.cell(row=5,column=i,value=dict["P_Name"]).fill=Fill
                else:
                    ws.cell(row=5,column=i).fill=Fill
    
    ##ws["B1"]="Hello Python"
    #if ws[az[1]+"1"] != "0:00" and ws[az[25]+"25"] != "0:00":
    #    time = "0:00"
    #    for i in range(1,26):
    #        for j in range:
    #            ws[az[i]+str(i)] = "0.00"
    #else:
        
        
        
#----------------------------------------------------------------
      
#------------------- save excel file --------------------------------      
    try:
        wb.save(SheetName)
    except PermissionError:
        print("Please, Close the Workbook before continuing")
#----------------------------------------------------------------

main("PyDSheet.xlsx")