from openpyxl import *
from openpyxl.styles import PatternFill
from datetime import datetime,time

import string
import random

class Graph():
    def __init__(self,Filename,Sheet):
        self.Filename = Filename
        self.Sheet = Sheet
        self.loading = [0,0] #(error_state,row_error)
        self.timer = [] #timestamp
        self.color = ["00FFFF00","00FF99CC","0099CC00","00FF9900","0000FFFF","00CC99FF","0033CCCC","00FF00FF","0099CCFF","00FFCC99","009999FF","00FFFFCC","0000CCFF","00FFCC00","000066CC"]
        self.head = [] #datetimestamp
        self.dict ={}
        
    def Poc(self):
    #------------------- open excel file --------------------------------
        try:
            self.wb = load_workbook(self.Filename) #use out
        except FileNotFoundError:
            #print("File not found")
            return self.loading #file not found send (0,0)
        else:
            self.loading[0] = 1
    #----------------------------------------------------------------

    #------------------- access sheet or create sheet --------------------------------    
        if "Graph Timing "+ self.Sheet in self.wb.sheetnames:
            self.ws = self.wb["Graph Timing "+self.Sheet]
        else:
            self.ws = self.wb.create_sheet("Graph Timing "+self.Sheet)
    #----------------------------------------------------------------
        self.ws.cell(row=1, column=1, value="Blending")
        self.ws.cell(row=1, column=2, value=self.Sheet)
        self.ws.cell(row=2, column=1, value="วันที่/เวลา")
        self.create_timeline()
        try:
            self.wsr = self.wb[self.Sheet]
        except KeyError:
            self.loading[0] = 3
            return self.loading
        else:
            targetcolumn = ["B","D"]
            self.create_header(targetcolumn)
            self.fill_cell()
        #print(self.head)
        
    #------------------- save excel file --------------------------------      
        if (self.loading[0] == 3 or self.loading[0] == 4):
            return self.loading
        else:
            try:
                self.wb.save(self.Filename)
            except PermissionError:
                #print("Please, Close the Workbook before continuing")
                return self.loading #send (1,0)
            else:
                self.loading[0] = 2
                return self.loading
    #----------------------------------------------------------------
  
    def create_timeline(self):
        start_hour = 0
        #stop_hour = 0
        #step_hour = 1
        start_minute = 0
        stop_minute = 31
        #step_minute = 30
        for hour in range(start_hour,24):
            for minute in range(start_minute,stop_minute+1,30):
                formatted_hour = f"{hour:02d}"
                formatted_minute = f"{minute:02d}"
                self.timer.append(f"{formatted_hour}:{formatted_minute}")
        self.timer.append(self.timer[0])
        insert_pos = 1
        round = len(self.timer)
        while True:
            if round == 0:
                break
            else:
                self.ws.cell(row=2, column=insert_pos+1, value=self.timer[insert_pos-1])
                insert_pos += 1
                round-=1
    
    def create_header(self,cols):
        for col in cols:
            for cell in self.wsr[col]:
                if cell.value == "วันที่เริ่มต้น" or cell.value == "วันที่เสร็จ" or cell.value in self.head:
                    continue
                else:
                    self.head.append(cell.value)
        insert_cell = 4
        loop = 0
        while True:
            if loop == len(self.head):
                break
            else:
                self.ws.cell(row=insert_cell, column=1, value=self.head[loop])
                loop += 1
                insert_cell += 2
    def fill_cell(self):
        #------------------- Ploting in sheet --------------------------------
        a = 0
        use_color = 0
        row_now = 0
        first_time = True
        #listdate = self.head
        for r in self.wsr.iter_rows(min_row=2):  # Start from the second row
            if first_time is True:
                row_now += 2
                first_time = False
            else:
                row_now += 1
            if use_color == len(self.color)-1:
                use_color = 0
            #use_color = random.randint(0,len(self.color)-1)
            round = 1
            #a = 0
            for cell in r:
                if cell.value is None:
                    break
                if round == 1:
                    self.dict["P_Name"] = cell.value
                elif round == 2:
                    self.dict["Date_Start"] = cell.value
                elif round == 3:
                    if type(cell.value) == type(time(10,0)):
                        self.dict["Time_Start"] = cell.value.strftime("%H:%M")
                    else:
                        self.dict["Time_Start"] = cell.value
                    if self.dict["Time_Start"] not in self.timer:
                        self.loading[0] = 3
                        self.loading[1] = row_now
                        return
                        #return self.loading #send (3,row_now)
                elif round == 4:
                    self.dict["Date_End"] = cell.value
                elif round == 5:
                    if type(cell.value) == type(time(10,0)):
                        self.dict["Time_End"] = cell.value.strftime("%H:%M")
                    else:
                        self.dict["Time_End"] = cell.value
                    if self.dict["Time_End"] not in self.timer:
                        self.loading[0] = 4
                        self.loading[1] = row_now
                        return
                        #return self.loading #send (4,row_now)
                else:
                    print("??")
                round+=1
            #print(self.dict)
            if len(self.dict) == 0:
                break
            else:
                start_hour = self.timer.index(self.dict["Time_Start"])+2
                end_hour = self.timer.index(self.dict["Time_End"])+3
                Fill = PatternFill(start_color=self.color[use_color],end_color=self.color[use_color],fill_type="solid")
                use_color+=1
                for a in range(0,len(self.head)):
                    if self.dict["Date_Start"] == self.head[a]:
                        for i in range(start_hour, end_hour):
                            if i == start_hour:
                                self.ws.cell(row=((2*a)+4),column=i,value=self.dict["P_Name"]).fill=Fill
                            else:
                                self.ws.cell(row=((2*a)+4),column=i).fill=Fill                                   
    #----------------------------------------------------------------   
                
    
#g = Graph("กราฟการทำงานของ ABB.xlsx","TestError")
#print(g.Poc())